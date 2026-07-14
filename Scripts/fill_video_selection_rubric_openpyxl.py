#!/usr/bin/env python3
"""
OpenPyXL version (preserves formulas/styles) to fill Video_Selection_Rubric_AutoScore.xlsx.

It updates ONLY the rubric input columns (E–N) and leaves Selection_Score/Include formulas intact.

Usage:
  python fill_video_selection_rubric_openpyxl.py \
      --rubric Video_Selection_Rubric_AutoScore.xlsx \
      --transcripts_dir transcripts \
      --out Video_Selection_Rubric_AutoScore_Filled.xlsx \
      --family_list Ransomware_Family_Coverage_List.xlsx
"""

import argparse
import re
from pathlib import Path
from collections import Counter
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook


TTP_KEYWORDS = [
    r"\bphish\w*\b", r"\bphishing\b", r"\bmalspam\b", r"\bvoice\s+phish\w*\b", r"\bvishing\b",
    r"\brdp\b", r"\bexposed rdp\b", r"\bbrute\s+force\b", r"\bpassword\s+spray\b",
    r"\bexploit\w*\b", r"\bvulnerab\w*\b", r"\b0-day\b",
    r"\bcobalt\s+strike\b", r"\bbeacon\b", r"\bmimikatz\b", r"\bkerberoast\w*\b",
    r"\bpsexec\b", r"\bwmi\b", r"\bwinrm\b",
    r"\brclone\b", r"\bmeganz\b", r"\bwinscp\b", r"\bftp\b", r"\bexfil\w*\b", r"\bdata\s+theft\b",
    r"\bdouble\s+extortion\b", r"\bleak\s+site\b",
    r"\bshadow\s+copy\b", r"\bvss\b", r"\bransom\s+note\b", r"\bencrypt\w*\b",
    r"\bscheduled\s+task\b", r"\bgpo\b", r"\bgroup\s+policy\b",
]

DFIR_CASE_STUDY_SIGNALS = [
    r"\bcase\s+summary\b", r"\bintrusion\b", r"\bwe\s+observed\b", r"\btimeline\b",
    r"\bbeach\s*head\b", r"\bday\s+\d+\b", r"\bwithin\s+\d+\s+(day|days|hour|hours)\b",
    r"\bincident\s+response\b", r"\bforensic\w*\b", r"\bpost[-\s]?incident\b",
]

PLATFORM_PATTERNS = {
    "ESXi": [r"\besxi\b", r"\bvmware\b", r"\bvcenter\b"],
    "Linux": [r"\blinux\b", r"\bubuntu\b", r"\bdebian\b", r"\bcentos\b", r"\brhel\b", r"\bred\s*hat\b"],
    "Windows": [r"\bwindows\b", r"\bactive\s+directory\b", r"\bdomain\s+controller\b", r"\blsass\b"],
}

DEPTH_HIGH = [
    r"\batt&ck\b", r"\bmitre\b", r"\btelemetry\b", r"\bioc(s)?\b", r"\byara\b", r"\bsigma\b",
    r"\bpcap\b", r"\bmemory\s+dump\b", r"\breverse\s+engineering\b", r"\bapi\b", r"\bprocess\s+injection\b",
    r"\blog(s)?\b", r"\bevidence\b", r"\bdetection\s+rule\b",
]
DEPTH_MED = [
    r"\bkill\s+chain\b", r"\blateral\s+movement\b", r"\bprivilege\s+escalation\b",
    r"\bexfiltration\b", r"\bpersistence\b", r"\bdiscovery\b",
]

RED_FLAG_PATTERNS = {
    "Marketing": [
        r"\b(contact\s+us|pricing|book\s+a\s+demo|request\s+a\s+demo)\b",
        r"\bour\s+product\b", r"\bwebinar\b", r"\bsponsored\b"
    ],
    "News": [
        r"\bbreaking\s+news\b", r"\bheadlines\b", r"\bnews\s+update\b"
    ],
    "LowSignal": [
        r"\b(what\s+is\s+ransomware|ransomware\s+explained)\b"
    ]
}


def normalize(text: str) -> str:
    text = text.lower()
    text = re.sub(r"\s+", " ", text)
    return text


def any_hit(text: str, patterns: List[str]) -> bool:
    return any(re.search(p, text, flags=re.IGNORECASE) for p in patterns)


def count_hits(text: str, patterns: List[str]) -> int:
    return sum(len(re.findall(p, text, flags=re.IGNORECASE)) for p in patterns)


def detect_platform(text: str) -> str:
    hits = {plat: count_hits(text, pats) for plat, pats in PLATFORM_PATTERNS.items()}
    chosen = [p for p, h in hits.items() if h > 0]
    if not chosen:
        return ""
    if len(chosen) == 1:
        return chosen[0]
    return "Mixed"


def detect_depth(text: str) -> str:
    high = count_hits(text, DEPTH_HIGH)
    med = count_hits(text, DEPTH_MED)
    if high >= 6:
        return "High"
    if high >= 2 or med >= 4:
        return "Medium"
    return "Low"


def detect_red_flags(text: str) -> str:
    flags = [label for label, pats in RED_FLAG_PATTERNS.items() if any_hit(text, pats)]
    return "/".join(flags)


def load_family_aliases(family_list_path: Path) -> Dict[str, str]:
    """
    Returns alias_lower -> canonical_family
    """
    import openpyxl
    wb = openpyxl.load_workbook(family_list_path)
    ws = wb.active
    # Find header columns
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c

    col_fam = headers.get("Ransomware_Family_Name")
    col_alias = headers.get("Alias_Names")
    if not col_fam:
        return {}

    alias_map: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        fam = ws.cell(row=r, column=col_fam).value
        if not fam:
            continue
        fam = str(fam).strip()
        alias_map[fam.lower()] = fam

        if col_alias:
            aliases = ws.cell(row=r, column=col_alias).value
            if aliases and str(aliases).strip() and str(aliases).strip() != "—":
                for a in [x.strip() for x in str(aliases).split(",")]:
                    if a:
                        alias_map[a.lower()] = fam
    return alias_map


def detect_specific_family(text: str, alias_map: Dict[str, str]) -> bool:
    if not alias_map:
        return False
    for alias_lower in alias_map.keys():
        alias_re = re.escape(alias_lower).replace(r"\ ", r"\s+")
        pat = rf"\b{alias_re}\b"
        if re.search(pat, text, flags=re.IGNORECASE):
            return True
    return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rubric", required=True)
    ap.add_argument("--transcripts_dir", default="transcripts")
    ap.add_argument("--out", default="Video_Selection_Rubric_AutoScore_Filled.xlsx")
    ap.add_argument("--family_list", default="")
    args = ap.parse_args()

    rubric_path = Path(args.rubric)
    transcripts_dir = Path(args.transcripts_dir)
    out_path = Path(args.out)

    alias_map: Dict[str, str] = {}
    if args.family_list:
        alias_map = load_family_aliases(Path(args.family_list))

    wb = load_workbook(rubric_path)
    ws = wb["Video_Selection_Rubric"] if "Video_Selection_Rubric" in wb.sheetnames else wb.active

    # Header mapping
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c

    def h(name: str) -> int:
        return headers[name]

    COL_VID = h("Video_ID")
    COL_FAM_MENTION = h("Ransomware_Family_Mentioned (Yes/No)")
    COL_SPEC_FAM = h("Specific_Family_Named (Yes/No)")
    COL_TTPS = h("TTPs_Mentioned (Yes/No)")
    COL_DFIR = h("DFIR_Case_Study (Yes/No)")
    COL_PLATFORM = h("Platform_Mentioned (Windows/Linux/ESXi)")
    COL_TRANSCRIPT = h("Transcript_Available (Yes/No)")
    COL_DEPTH = h("Estimated_Technical_Depth (Low/Medium/High)")
    COL_REDFLAG = h("Red_Flags (News/Marketing/LowSignal)")

    def set_if_empty(r: int, c: int, value: str):
        cur = ws.cell(row=r, column=c).value
        if cur is None or str(cur).strip() == "":
            ws.cell(row=r, column=c, value=value)

    # Data starts at row 3 in your template (row 2 is guidance)
    for r in range(3, ws.max_row + 1):
        vid = ws.cell(row=r, column=COL_VID).value
        if not vid:
            continue
        vid = str(vid).strip()
        if not vid or vid.lower().startswith("scoring guidance"):
            continue

        tfile = transcripts_dir / f"{vid}.txt"
        if not tfile.exists():
            set_if_empty(r, COL_TRANSCRIPT, "No")
            continue

        text = normalize(tfile.read_text(encoding="utf-8", errors="ignore"))

        set_if_empty(r, COL_TRANSCRIPT, "Yes")

        has_specific = detect_specific_family(text, alias_map) if alias_map else False
        if has_specific:
            set_if_empty(r, COL_SPEC_FAM, "Yes")
            set_if_empty(r, COL_FAM_MENTION, "Yes")
        else:
            has_ransomware_word = bool(re.search(r"\bransomware\b", text, flags=re.IGNORECASE))
            set_if_empty(r, COL_FAM_MENTION, "Yes" if has_ransomware_word else "No")
            set_if_empty(r, COL_SPEC_FAM, "No")

        set_if_empty(r, COL_TTPS, "Yes" if any_hit(text, TTP_KEYWORDS) else "No")
        set_if_empty(r, COL_DFIR, "Yes" if any_hit(text, DFIR_CASE_STUDY_SIGNALS) else "No")

        platform = detect_platform(text)
        if platform:
            set_if_empty(r, COL_PLATFORM, platform)

        set_if_empty(r, COL_DEPTH, detect_depth(text))

        flags = detect_red_flags(text)
        if flags:
            set_if_empty(r, COL_REDFLAG, flags)

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
